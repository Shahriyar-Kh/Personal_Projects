from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image,ImageTk
import os
from  tkinter.ttk import Combobox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib
  
framebg="#EDEDED"
framefg="#06383D"

root=Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg="#06283D")

#Create sheet file with columns


file=pathlib.Path(r'E:\Python_Prac\Ch17_GUI_Tkinter\\1_Student_Registration_sys\Student_data.xlsx')
if file.exists(): 
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Registration No."
    sheet["B1"]="Name"
    sheet["C1"]="Class"
    sheet["D1"]="Gender"
    sheet["E1"]="DOB"
    sheet["F1"]="Date of Regitration"
    sheet["G1"]="Religion"
    sheet["H1"]="Skill"
    sheet["I1"]="Father Name"
    sheet["J1"]="Mother Name"
    sheet["K1"]="Father's Occupation"
    sheet["L1"]="Mother's Occupation"
    file.save(r'E:\Python_Prac\Ch17_GUI_Tkinter\\1_Student_Registration_sys\Student_data.xlsx')
    # file.close()

def Exit():
    btn_Exit.destroy()

#************************************  Uplaod the profile pecture *******************************************
def showimage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),title="Select Title Image",filetypes=(("JPG Files","*.jpg"),("PNG Files","*.png"),("JPEG Files","*.jpeg"),("All Files","*.txt")))

    img=Image.open(filename)
    resized_img=img.resize((190,190))
    photo1=ImageTk.PhotoImage(resized_img)
    upload_l.config(image=photo1)
    upload_l.image=photo1

#************************************ Registration no ************************************
#it created to automatic enter registration no
def registration_no():
    file=openpyxl.load_workbook(r"E:\Python_Prac\Ch17_GUI_Tkinter\\1_Student_Registration_sys\Student_data.xlsx")
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value
    try:
        Registrationvar.set(max_row_value+1)
    except:
        Registrationvar.set("1")
    
    # file.close()


def Clear():
    global img
    Namevar.set("")
    dobvar.set("")
    religionvar.set("")
    skillsvar.set("")
    fnamevar.set("")
    Mnamevar.set("")
    F_Occupation.set("")
    M_Occupation.set("")
    class_cambo.set("Select Class")

    registration_no()
    btn_Save.config(state="normal")
    img1=PhotoImage(file=r"E:\Python_Prac\Ch17_GUI_Tkinter\\1_Student_Registration_sys\upload3.png")
    upload_l.config(image=img1)
    upload_l.image=img1
    img=""



def Save_data():
    regno=Registrationvar.get()
    name=Namevar.get()
    cls=class_cambo.get()
    try:
        gender1=radiovar.get()
    except:
        messagebox.showerror("Error","Select Gender")

    dob=dobvar.get()
    date1=Datevar.get()
    fname=fnamevar.get()
    mname=Mnamevar.get()
    f_ocp=F_Occupation.get()
    m_ocp=M_Occupation.get()
    religion=religionvar.get()
    skills=skillsvar.get()


    if name=="" or cls=="Selected Class" or gender1=="" or dob=="" or religion=="" or fname=="" or mname=="" or f_ocp=="" or m_ocp=="":
        messagebox.showerror("Error","Few Data is missing!!!")
    else:
        file=openpyxl.load_workbook(r'E:\Python_Prac\Ch17_GUI_Tkinter\\1_Student_Registration_sys\Student_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=regno)
        sheet.cell(column=2,row=sheet.max_row,value=name)
        sheet.cell(column=3,row=sheet.max_row,value=cls)
        sheet.cell(column=4,row=sheet.max_row,value=gender1)
        sheet.cell(column=5,row=sheet.max_row,value=dob)
        sheet.cell(column=6,row=sheet.max_row,value=date1)
        sheet.cell(column=7,row=sheet.max_row,value=religion)
        sheet.cell(column=8,row=sheet.max_row,value=skills)
        sheet.cell(column=9,row=sheet.max_row,value=fname)
        sheet.cell(column=10,row=sheet.max_row,value=mname)
        sheet.cell(column=11,row=sheet.max_row,value=f_ocp)
        sheet.cell(column=12,row=sheet.max_row,value=m_ocp)
        
        file.save(r'E:\Python_Prac\Ch17_GUI_Tkinter\\1_Student_Registration_sys\Student_data.xlsx')

        try:
            img.save(rf"E:\Python_Prac\Ch17_GUI_Tkinter\\1_Student_Registration_sys\SRS_Pics\{name}"+str(regno)+".jpg")
        except:
            messagebox.showinfo("Info","Profile Picture is not available!!!!")
        messagebox.showinfo("info","Successfully data entered!!!")

        Clear()#clear the felds
        registration_no()#it will recheck reg no and reissue now





    


def Search():
    get_text=Searchvar.get() #Taking input from entry box
    Clear()#Clear all data already avalible in entry box

    btn_Save.config(state="disable") #after clicking on search

    file=openpyxl.load_workbook(r'E:\Python_Prac\Ch17_GUI_Tkinter\\1_Student_Registration_sys\Student_data.xlsx')
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==int(get_text):
            name=row[0]

            regno_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]
            print(regno_position)
            print(reg_number)

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid registration number")
    
    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value

    Registrationvar.set(x1)
    Namevar.set(x2)
    class_cambo.set(x3)
    if x4=="Female":
        r2.select()
    else:
        r1.select()

    dobvar.set(x5)
    Datevar.set(x6)
    religionvar.set(x7)
    skillsvar.set(x8)
    fnamevar.set(x9)
    Mnamevar.set(x10)
    F_Occupation.set(x11)
    M_Occupation.set(x12)

    img=Image.open(rf"E:\Python_Prac\Ch17_GUI_Tkinter\\1_Student_Registration_sys\SRS_Pics\{x2}{str(x1)}.jpg")
    resized_img=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_img)
    upload_l.config(image=photo2)
    upload_l.image=photo2


def Update():
    regno = Registrationvar.get()
    name = Namevar.get()
    cls = class_cambo.get()
    try:
        gender1 = radiovar.get()
    except:
        messagebox.showerror("Error", "Select Gender")

    dob = dobvar.get()
    date1 = Datevar.get()
    fname = fnamevar.get()
    mname = Mnamevar.get()
    f_ocp = F_Occupation.get()
    m_ocp = M_Occupation.get()
    religion = religionvar.get()
    skills = skillsvar.get()

    try:
        file = openpyxl.load_workbook(r'E:\Python_Prac\Ch17_GUI_Tkinter\1_Student_Registration_sys\Student_data.xlsx')
        sheet = file.active

        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row[0] == regno:
                reg_number = idx
                break

        sheet.cell(row=reg_number, column=2, value=name)
        sheet.cell(row=reg_number, column=3, value=cls)
        sheet.cell(row=reg_number, column=4, value=gender1)
        sheet.cell(row=reg_number, column=5, value=dob)
        sheet.cell(row=reg_number, column=6, value=date1)
        sheet.cell(row=reg_number, column=7, value=religion)
        sheet.cell(row=reg_number, column=8, value=skills)
        sheet.cell(row=reg_number, column=9, value=fname)
        sheet.cell(row=reg_number, column=10, value=mname)
        sheet.cell(row=reg_number, column=11, value=f_ocp)
        sheet.cell(row=reg_number, column=12, value=m_ocp)

        file.save(r'E:\Python_Prac\Ch17_GUI_Tkinter\1_Student_Registration_sys\Student_data.xlsx')

        try:
            img.save(rf"E:\Python_Prac\Ch17_GUI_Tkinter\1_Student_Registration_sys\SRS_Pics\{name}{regno}.jpg")
        except:
            messagebox.showinfo("Update", "Profile Picture is not available!!!!")
        messagebox.showinfo("Update", "Successfully data entered!!!")

        Clear()  # clear the fields
    except Exception as e:
        print(e)

#Fram 1
Label(root,text="Email: shahriyarkhanpk1@gmail.com",width=10,height=3,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
#frame2
Label(root,text="STUDENT REGISTRATION",width=10,height=2,bg="#C36464",fg="#fff",font="arial 20 bold").pack(side=TOP,fill=X)

#Search to update
Searchvar=StringVar()
Search_entry=Entry(root,textvariable=Searchvar,width=15,bd=2,font="Arial 20 ")
#Bind the resizing behavior directly to the window's <Configure> event
def on_resize(event):
    new_x = root.winfo_width() - 450  # Adjust as needed
    Search_entry.place(x=new_x, y=70)
    Search_btn.place(x=new_x+240,y=70)

root.bind("<Configure>", on_resize)
#Search icon
imageicon=PhotoImage(file='E:\Python_Prac\Ch17_GUI_Tkinter\\1_Student_Registration_sys\search2.png')
Search_btn=Button(root,text="Search",compound="left",image=imageicon,bg="#68ddfa",font="arial 13 bold",width=123,command=Search)
Search_btn.place(x=1060,y=70)

#update image
updateicon=PhotoImage(file=r"E:\Python_Prac\Ch17_GUI_Tkinter\\1_Student_Registration_sys\update.png")
update_btn=Button(root,text="Update",font="Arial 13 bold",command=Update,compound=LEFT,image=updateicon,bg="#c36464",)
update_btn.place(x=110,y=64)

#Registration and data 
Label(root,text="Registration No:",font="Arial 13",fg=framefg,bg=framebg).place(x=30,y=150)
Label(root,text="Date:",font="Arial 13",fg=framefg,bg=framebg).place(x=500,y=150)

Registrationvar=IntVar()
Datevar=StringVar()
reg_entry=Entry(root,textvariable=Registrationvar,width=15,font="arial 14")
reg_entry.place(x=160,y=150)

registration_no()

#Date
today=date.today()
d1=today.strftime(r"%d/%m/%Y")
date_entry=Entry(root,textvariable=Datevar,width=15,font="arial 14")
date_entry.place(x=550,y=150)
Datevar.set(d1)

#Student details
obj1=LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj1.place(x=30,y=200)
Label(obj1,text="Full Name:",font="Arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj1,text="Date of Birth:",font="Arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj1,text="Gender:",font="Arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

Label(obj1,text="Class:",font="Arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj1,text="Religion:",font="Arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj1,text="Skills:",font="Arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

Namevar=StringVar()
Name_entry=Entry(obj1,textvariable=Namevar,font="arial 14")
Name_entry.place(x=160,y=50)

dobvar=StringVar()
dob_entry=Entry(obj1,textvariable=dobvar,font="arial 14")
dob_entry.place(x=160,y=100)

radiovar=StringVar()
radiovar.set("radio")
r1=Radiobutton(obj1,text="Male",variable=radiovar,value="Male",fg=framefg,bg=framebg)
r1.place(x=150,y=150)
r2=Radiobutton(obj1,text="Female",variable=radiovar,value="Female",fg=framefg,bg=framebg)
r2.place(x=200,y=150)
class_cambo=Combobox(obj1,value=["1","2","3","4","5","6","7","8","9","10","11","12","BS","MS"],font="Roboto 14",width=17,state="r")
class_cambo.place(x=630,y=50)
class_cambo.set("Select Class")

religionvar=StringVar()
religion_entry=Entry(obj1,textvariable=religionvar,font="arial 14")
religion_entry.place(x=630,y=100)

skillsvar=StringVar()
skill_entry=Entry(obj1,textvariable=skillsvar,font="arial 14")
skill_entry.place(x=630,y=150)

#Parent details
obj2=LabelFrame(root,text="Parent's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Father's Name",font="arial 13",fg=framefg,bg=framebg).place(x=30,y=50)
Label(obj2,text="Occupation",font="arial 13",fg=framefg,bg=framebg).place(x=30,y=100)

fnamevar=StringVar()
fname_entry=Entry(obj2,textvariable=fnamevar,font="arial 14")
fname_entry.place(x=150,y=50)

F_Occupation=StringVar()
f_Ocp_entry=Entry(obj2,textvariable=F_Occupation,font="arial 14")
f_Ocp_entry.place(x=150,y=100)

Label(obj2,text="Mother's Name",font="arial 13",fg=framefg,bg=framebg).place(x=500,y=50)
Label(obj2,text="Occupation",font="arial 13",fg=framefg,bg=framebg).place(x=500,y=100)

Mnamevar=StringVar()
M_name_entry=Entry(obj2,textvariable=Mnamevar,font="arial 14")
M_name_entry.place(x=640,y=50)

M_Occupation=StringVar()
M_Ocp_entry=Entry(obj2,textvariable=M_Occupation,font="arial 14")
M_Ocp_entry.place(x=640,y=100)


#Upload image

f=Frame(root,bd=3,bg="black",width=200,height=200,relief=SUNKEN)
f.place(x=1000,y=150)

uploadimg=PhotoImage(file=r"E:\Python_Prac\Ch17_GUI_Tkinter\\1_Student_Registration_sys\upload3.png")
upload_l=Label(f,bg="black",image=uploadimg)
upload_l.place(x=0,y=0)


#Buttons

btn_upload=Button(root,text="Upload",width=19,height=2,font="Arial 13 bold",bg="lightblue",command=showimage)
btn_upload.place(x=1000,y=370)

btn_Save=Button(root,text="Save",width=19,height=2,font="Arial 13 bold",bg="lightgreen",command=Save_data)
btn_Save.place(x=1000,y=450)

btn_Reset=Button(root,text="Reset",width=19,height=2,font="Arial 13 bold",bg="lightpink",command=Clear)
btn_Reset.place(x=1000,y=530)

btn_Exit=Button(root,text="Exit",width=19,height=2,font="Arial 13 bold",bg="grey",command=Exit)
btn_Exit.place(x=1000,y=610)
root.mainloop()